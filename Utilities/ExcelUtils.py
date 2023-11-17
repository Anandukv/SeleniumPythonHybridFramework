import openpyxl


def getrowcount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)


def getcolumncount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)


def readdata(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnnum).value


def writedata(file, sheetName, rownum, columnnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)


def get_data_from_excel(file, sheetName):
    final_list = []
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(2, total_rows ):
        row_list = []
        for c in range(1, total_cols ):
            row_list.append(sheet.cell(row=r, column=c).value)
        final_list.append(row_list)

    return final_list
